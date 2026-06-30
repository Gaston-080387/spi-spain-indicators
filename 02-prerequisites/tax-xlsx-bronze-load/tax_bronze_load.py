"""
Sprint 2 (part II) - Tax AEAT (Local Validation).
Generate a script to load data from an URL.
Get a XLSX file and build a DataFrame.
Download the data in Parquet format to the working directory.
"""

import logging
from datetime import datetime, timezone
from pathlib import Path
import requests

from openpyxl import load_workbook
import pandas as pd

# ────────────────────────────────────────────────────────────────────────────
# Configure Logging
# ────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("tax_xlsx_bronze_load")

# ────────────────────────────────────────────────────────────────────────────
# Configure environment variables
# ────────────────────────────────────────────────────────────────────────────
URL_XLSX = (
    "https://sede.agenciatributaria.gob.es/static_files/Sede/Tema/Estadisticas/"
    "Recaudacion_Tributaria/Informes_mensuales/Ingresos_por_Delegaciones.xlsx"
)
SHEET_NAME = "datos_delegaciones"
INPUT_FOLDER = (
    "/Users/gastonbaloira/Projects/Portfolio/01_spi-indicators/"
    "02-prerequisites/tax-xlsx-bronze-load/input/"
)
OUTPUT_FOLDER = (
    "/Users/gastonbaloira/Projects/Portfolio/01_spi-indicators/"
    "02-prerequisites/tax-xlsx-bronze-load/output/"
)

# ────────────────────────────────────────────────────────────────────────────
# Create function download_file
# ────────────────────────────────────────────────────────────────────────────
def download_file() -> Path:
    """
    Download the direct Path to the input folder of the working directory.
    Returns:
        Path: Local path to the downloaded XLSX file.    
    """
    file_name = URL_XLSX.split("/")[-1]
    input_path_folder = Path(INPUT_FOLDER) / file_name
    log.info("Downloading: %s", file_name)
    response = requests.get(URL_XLSX)
    response.raise_for_status()
    input_path_folder.write_bytes(response.content)
    log.info(
        "File downloaded: %s  |  Bytes: %s",
        file_name, input_path_folder.stat().st_size,
    )
    return input_path_folder

# ────────────────────────────────────────────────────────────────────────────
# Create function cell_to_str
# ────────────────────────────────────────────────────────────────────────────
def cell_to_str(value) -> str:
    """Convert a openpyxl cell to string. (Bronze layer)"""
    if value is None: 
        return ''
    return str(value)

# ────────────────────────────────────────────────────────────────────────────
# Create function parse_tax_xlsx
# ────────────────────────────────────────────────────────────────────────────
def parse_tax_xlsx(file_path: Path) -> pd.DataFrame:
    """
    Read the XLSX file from the input folder and return a DataFrame
    Args:
        file_path: Path from download_file function.
    Returns:
        pd.DataFrame: A Dataframe with data and 2 lineage columns 
        (_ingestion_timestamp & _source_file)
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sh = wb[SHEET_NAME]
    all_rows = []
    for row in sh.iter_rows(values_only=True):
        all_rows.append([cell_to_str(v) for v in row])
    col_names = [f"col_{i:02d}" for i in range(1, len(all_rows[0]) + 1)]
    df = pd.DataFrame(all_rows, columns=col_names)
    df["_ingestion_timestamp"] = datetime.now(timezone.utc).isoformat()
    df["_source_file"] = file_path.name
    log.info(
        "Parsed: %s  |  Rows: %s  |  Cols: %s",
        file_path.name, df.shape[0], df.shape[1],
    )
    return df 

# ────────────────────────────────────────────────────────────────────────────
# Create function main to run the entire process
# ────────────────────────────────────────────────────────────────────────────
def main() -> None:
    """
    Main function that executes the entire process Data loading from AEAT URL.
    """
    path = download_file()
    df = parse_tax_xlsx(path)
    output_path = Path(OUTPUT_FOLDER) / f"{path.stem}.parquet"
    df.to_parquet(output_path, index=False)
    log.info("Stored: %s", output_path)

if __name__ == "__main__":
    main()